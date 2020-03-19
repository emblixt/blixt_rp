import pandas as pd
import os
import re
import numpy as np
from datetime import datetime
import logging

from utils.utils import isnan

logger = logging.getLogger(__name__)


def project_wells(filename, working_dir):
    table = pd.read_excel(filename, header=1, sheet_name='Wells table')
    result = {}
    for i, ans in enumerate(table['Use']):
        if ans == 'Yes':
            temp_dict = {}
            log_dict = {}
            for key in list(table.keys()):
                if (key == 'las file') or (key == 'Use'):
                    continue
                elif (key == 'Given well name') or (key == 'Note'):
                    if isinstance(table[key][i], str):
                        if key == 'Given well name':
                            value = fix_well_name(table[key][i])
                        else:
                            value = table[key][i]
                        temp_dict[key] = value
                    else:
                        temp_dict[key] = None  # avoid NaN
                else:
                    val = table[key][i]
                    if isnan(val):
                        continue
                    else:
                        this_list = make_clean_list(table[key][i], small_cap=True)
                        for log_name in this_list:
                            log_dict[log_name] = key
                temp_dict['logs'] = log_dict
            temp_file = test_file_path(table['las file'][i], working_dir)
            if temp_file is False:
                warn_txt = 'Warning, las file {} does not exist'.format(table['las file'][i])
                logger.warning(warn_txt)
                raise Warning(warn_txt)
            result[test_file_path(table['las file'][i], working_dir)] = temp_dict
    return result


def harmonize_wells():
    # TODO
    # write a function that harmonizes the different log names in a set of las files
    # i.e. so that shale volume has the same log name in all wells.
    # This could be done on input without modifying the las files
    pass


def collect_project_wells(well_table, target_dir):
    """
    Copies all las files in the well table (the output from project_wells()) to the folder target_dir.

    :param well_table:
    :return:
    """
    from shutil import copyfile
    for las_file in list(well_table.keys()):
        if os.path.isfile(las_file):
            short_name = os.path.split(las_file)[-1]
            print('Copying file {} to {}'.format(
                short_name, target_dir
            ))
            copyfile(las_file, os.path.join(target_dir, short_name))


def read_sums_and_averages(filename, header=20):
    table = pd.read_excel(filename, header=header)
    unique_layers = unique_names(table, 'Name', well_names=False)
    answer = {}
    for layer in unique_layers:
        answer[layer] = {}

    for key in list(table.keys()):
        if key == 'Name':
            continue
        for i, value in enumerate(table[key]):
            answer[table['Name'][i]][key] = value

    return answer


def write_sums_and_averages(filename, line_of_data):
    # This function creates xlsx files. so please use Excel to save them as
    # xls files before attempting to load them into RokDoc
    # Additional columns are added after the  'ShaleVolumeAspectRatio' column, which are not
    # read by RokDoc
    if filename.split('.')[-1] == 'xls':
        filename += 'x'

    from openpyxl import load_workbook, Workbook
    if not os.path.isfile(filename):
        print('Creating new RokDoc Sums and Averages file')
        newfile = True
        wb = Workbook()
    else:
        print('Appending to existing RokDoc Sums and averages file')
        newfile = False
        wb = load_workbook(filename)

    ws = wb.active
    if newfile:
        ws.append(['Averages Set output from simple python script well_tops.py on {}'.format(
            datetime.now().isoformat())])
        ws.append(['Template Version: 1'])
        ws.append(['Depth units:             m'])
        ws.append(['Time units:              ms'])
        ws.append(['Velocity units:          m/s'])
        ws.append(['Density units:           g/cm3'])
        ws.append(['Porosity units:          fract'])
        ws.append(['AI units:                g/cm3_m/s'])
        ws.append(['SI units:                g/cm3_m/s'])
        ws.append(['M units:                 GPa'])
        ws.append(['MU units:                GPa'])
        ws.append(['K (Bulk Modulus) units:  GPa'])
        ws.append(['Lambda units:            GPa'])
        ws.append(['E units:                 GPa'])
        ws.append(['Lambda Mu units:         fract'])
        ws.append(['Mu Rho units:            GPa_g/cm3'])
        ws.append(['Lambda Rho units:        GPa_g/cm3'])
        ws.append(['Saturation units:        fract'])
        ws.append(['Volume units:            fract'])
        ws.append(['TableStart:'])
        ws.append(
            [
                'Name', 'Well', 'ZType', 'TopDepth', 'BaseDepth', 'MidPointDepth',
                'VpMean', 'VsMean', 'RhoMean', 'VpMedian', 'VsMedian', 'RhoMedian',
                'VpMode', 'VsMode', 'RhoMode', 'PorosityType', 'PorosityMean',
                'PorosityStdDev', 'Net', 'NetToGross', 'EpsilonMean', 'DeltaMean',
                'GammaMean', 'EpsilonMedian', 'DeltaMedian', 'GammaMedian',
                'EpsilonMode', 'DeltaMode', 'GammaMode', 'VpStdDev', 'VsStdDev',
                'RhoStdDev', 'EpsilonStdDev', 'DeltaStdDev', 'GammaStdDev',
                'VpVsCorrCoef', 'VpRhoCorrCoef', 'VsRhoCorrCoef', 'AI', 'SI',
                'M', 'MU', 'KBulkModulus', 'PR', 'Lambda', 'E', 'LambdaMu',
                'MuRho', 'LambdaRho', 'ShaleVolumeMean', 'ShaleVolumeStdDev',
                'ShaleVolumeInclusionShape', 'ShaleVolumeAspectRatio', 'Classification',
                'DateAdded'
            ]
        )
    ws.append(line_of_data)
    wb.save(filename)
    wb.close()


def read_tops(filename, top=True, zstick='md', frmt=None):
    if frmt == 'petrel':
        return read_petrel_tops(filename, top=top, zstick=zstick)
    elif frmt == 'npd':
        return read_npd_tops(filename, top=top, zstick=zstick)
    elif frmt == 'rokdoc':
        return read_rokdoc_tops(filename, top=top, zstick=zstick)
    else:
        raise IOError('No tops for format {}'.format(frmt))


def read_rokdoc_tops(filename, header=4, top=True, zstick='md'):
    """
    :param top:
        bool
        if True, the top of each marker/top is returned
        if False, not implemented

    :param zstick:
        str
        adapted after RokDoc.
        Can be:
            'md', 'tvdkb','twt', 'tvdss',

    """
    if not top:
        raise NotImplementedError('Only top of markers / tops are available')

    if zstick == 'md':
        key_name = 'MD'
    elif zstick == 'tvdkb':
        key_name = 'TVDkb'
    elif zstick == 'twt':
        key_name = 'TWT'
    elif zstick == 'tvdss':
        key_name = 'TVDss'
    else:
        key_name = None
        NotImplementedError('ZSTICK = {} is not implemented'.format(zstick))

    tops = pd.read_excel(filename, header=header)
    return return_dict_from_tops(tops, 'Well Name', 'Horizon', key_name)


def read_npd_tops(filename, header=None, top=True, zstick='md'):
    if zstick != 'md':
        NotImplementedError('ZSTICK = {} is not implemented'.format(zstick))
    if top:
        key_name = 'Top depth [m]'
    else:
        key_name = 'Bottom depth [m]'

    tops = pd.read_excel(filename)
    return return_dict_from_tops(tops, 'Wellbore name', 'Lithostrat. unit', key_name)


def read_petrel_tops(filename, header=None, top=True, zstick='md'):
    if zstick != 'md':
        NotImplementedError('ZSTICK = {} is not implemented'.format(zstick))
    if not top:
        NotImplementedError('Only MD is implemented for Petrel top files')
        key_name = None
    else:
        key_name = 'MD'

    tops = pd.read_excel(filename)
    return return_dict_from_tops(tops, 'Well identifier', 'Surface', key_name)


def test_file_path(file_path, working_dir):
    if os.path.isfile(file_path):
        return file_path
    elif os.path.isfile(os.path.join(working_dir, file_path)):
        return os.path.join(working_dir, file_path)
    else:
        return False


def fix_well_name(well_name):
    if isinstance(well_name, str):
        return well_name.replace('/', '_').replace('-', '_').replace(' ', '').upper()
    else:
        return


def unique_names(table, column_name, well_names=True):
    """
    :param table:
        panda object
        as returned from pandas.read_excel()
    returns the list of unique values in a column named 'column_name'
    """
    if well_names:
        return [fix_well_name(x) for x in list(set(table[column_name])) if isinstance(x, str)]
    else:
        return [x for x in list(set(table[column_name])) if isinstance(x, str)]


def return_dict_from_tops(tops, well_key, top_key, key_name):
    unique_wells = unique_names(tops, well_key)
    answer = {}

    for well_name in unique_wells:
        answer[well_name] = {}

    for well_name in unique_wells:
        for i, marker_name in enumerate(list(tops[top_key])):
            if fix_well_name(tops[well_key][i]) != well_name:
                continue  # not on the right well
            answer[well_name][marker_name.upper()] = tops[key_name][i]

    return answer


def make_clean_list(input_str, small_cap=False):
    if not isinstance(input_str, str):
        raise IOError('Only accept strings')
    if small_cap:
        return [x.strip().lower() for x in input_str.split(',') if x.strip() != '']
    else:
        return [x.strip() for x in input_str.split(',') if x.strip() != '']


def write_las(filename, wh, lh, data, overwrite=False):
    """
    Write the well data to a las format file.

    :param filename:
        str
        full path name to las file
    :param wh:
        core.well.Header
        well header =  w.header, where w is a core.well.Well object
    :param lh:
        core.well.Header
        log header = w.log_blocks['LogBlock name'].header, where w is a core.well.Well object
    :param data:
        dict
        data = w.log_blocks['LogBlock name'].logs, where w is a core.well.Well object
    :param overwrite:
        bool
        Set to True to allow overwriting an existing las file
    :return:
    """
    if os.path.isfile(filename) and (not overwrite):
        warn_txt = 'File {} already exist. Write cancelled'.format(filename)
        print('WARNING: {}'.format(warn_txt))
        logger.warning(warn_txt)
        return

    out = (
        '#----------------------------------------------------------------------------\n'
        '~VERSION INFORMATION\n'
        'VERS.            2.0                  :CWLS LOG ASCII STANDARD -VERSION 2.0\n'
        'WRAP.            NO                   :ONE LINE PER DEPTH STEP\n'
        '#\n'
    )

    out += '# {}\n'.format(wh['creation_info'].value)
    out += '# NOTE: {}\n'.format(wh['note'].value)
    out += '# Written to las on: {}\n'.format(datetime.now().isoformat())
    out += '# Modified on: {}\n'.format(wh['modification_date'].value)
    for key, value in data.items():
        if value.header['modification_history'] is not None:
            out += '#  Modification: {}: {}\n'.format(key, value.header['modification_history'])

    out += (
        '#--------------------------------------------------------------------\n'
        '~WELL INFORMATION\n'
        '#MNEM .UNIT      DATA                 :DESCRIPTION OF MNEMONIC\n'
        '#----------      ------------         -------------------------------\n'
    )

    # add info about start stop etc. from LogBlock header
    for key in list(lh.keys()):
        if key in ['name', 'creation_info', 'creation_date', 'modification_date']:
            continue
        out += '{0: <7}.{1: <9}{2: <21}:{3:}\n'.format(
            key.upper(),
            lh[key].unit,
            str(lh[key].value) if lh[key].value is not None else '',
            lh[key].desc.upper()
        )

    # add info well header
    for key in list(wh.keys()):
        if key in ['name', 'note', 'creation_info', 'creation_date', 'modification_date']:
            continue
        out += '{0: <7}.{1: <9}{2: <21}:{3:}\n'.format(
            key.upper(),
            wh[key].unit,
            str(wh[key].value) if wh[key].value is not None else '',
            wh[key].desc.upper()
        )

    out += (
        '#\n'
        '# ----------------------------------------------------------------------------\n'
        '~CURVE INFORMATION\n'
        '# MNEM.UNIT                                         : CURVE DESCRIPTION\n'
        '# ----------                                        -------------------------------\n'
    )
    i = 1
    out += '{0: <20}.{1: <33}: {2: <9}{3:}\n'.format(
        'DEPTH',
        data['depth'].header['unit'],
        i,
        data['depth'].header['desc']
    )
    for key, value in data.items():
        if key == 'depth':
            continue
        i += 1
        out += '{0: <20}.{1: <33}: {2: <9}{3:}\n'.format(
            key.upper(),
            value.header['unit'],
            i,
            value.header['desc']
        )
    out += (
        '#\n'
        '# ----------------------------------------------------------------------------\n'
        '~A                  '
    )

    # write data column headers
    for key in list(data.keys()):
        if key == 'depth':
            continue
        out += '{0: <20}'.format(key.upper())
    out += '\n'

    # start writing data
    for i, md in enumerate(data['depth'].data):
        out += '{0: <20}'.format(md)
        for key in list(data.keys()):
            if key == 'depth':
                continue
            out += '{0: <20}'.format(
                wh['null'].value if np.isnan(data[key].data[i]) else data[key].data[i]
            )
        out += '\n'


    with open(filename, 'w+') as f:
        f.write(out)


def convert(lines, file_format='las'):
    """
    class handling wells, with logs, and well related information
    The reading .las files is more or less copied from converter.py
        https://pypi.org/project/las-converter/
    """
    def parse(x):
        try:
            x = int(x)
        except ValueError:
            try:
                x = float(x)
            except ValueError:
                pass
        return x

    def rename_depth(key):
        """
        Helper function that translates different "depth" names to a common "depth" name.
        :param key:
            str
        :return:
            str
        """
        return_name = 'depth'
        if key.lower() == 'dept':
            return return_name
        else:
            return key

    def get_current_section(line):
        if '~V' in line : return 'version'
        if '~W' in line: return 'well_info'
        if '~C' in line: return 'curve'
        if '~P' in line: return 'parameter'
        if '~O' in line: return 'other'
        if '~A' in line: return 'data'
        # ~ Unregistered section
        return None

    def add_section(well_dict, section, mnem, content):
        if section == "data":
            if isinstance(content, list):
                well_dict[section][mnem] = content
            else:
                well_dict[section][mnem].append(content)
        elif section == "other":
            well_dict[section] += "".join(str(content).strip())
        else:
            well_dict[section][mnem] = content
        return well_dict

    generated_keys = []
    null_val = None
    section = ""
    rules = {"version", "well_info", "parameter", "curve"}
    descriptions = []
    curve_names = None
    well_dict = {"version": {}, "well_info": {}, "curve": {}, "parameter": {}, "data": {}, "other": ""}
    if file_format == 'RP well table':
        null_val = 'NaN'

    for line in lines:
        content = {}

        if isinstance(line, bytes):
            line = line.decode("utf-8").strip()

        # line just enter or "\n"
        if len(line) <= 1: continue
        # comment
        if "#" in line: continue

        # section
        if "~" in line:
            section = get_current_section(line)

            # get section version first
            if section == "version":
                continue

            # generate keys of log[data] based on log[curve]
            if section == "data":
                generated_keys = [e.lower() for e in well_dict["curve"].keys()]
                for key in generated_keys:
                    key = rename_depth(key)
                    # inital all key to empty list
                    well_dict = add_section(well_dict, section, key, [])

            continue

        if file_format == 'RP well table':
            if line[:7] == 'Columns':
                section = 'RP header'
                continue  # jump into header

            if section == 'RP header':
                if line[2:5] == ' - ':
                    descriptions.append(line.split(' - ')[-1].strip())

            if line[:7] == 'Well ID':
                # parse curve names
                curve_names = [t.strip().lower() for t in line.split('\t')]
                section = 'dummy_value'

            if line[:4] == '  No':
                # parse line of units
                # unit_names = [t.strip() for t in line.split('\t')]
                unit_names = [t.strip() for t in line.split()]
                unit_names = [t.replace('[', '').replace(']', '') for t in unit_names]

                for this_curve_name, this_unit_name, this_description in zip(curve_names, unit_names, descriptions):
                    well_dict = add_section(well_dict, 'curve',
                                            this_curve_name,
                                            {'api_code': None, 'unit': this_unit_name, 'desc': this_description}
                                            )
                generated_keys = [key for key in curve_names]
                section = 'data'
                # initiate all key to empty list
                for key in generated_keys:
                    # TODO This will probably fail. Test and compare with 'original' converter
                    well_dict.add_section(well_dict, key, [])
                continue  # jump into data

        # unregistered section
        if section is None: continue

        if section in rules:
            # index of seperator
            if re.search("[.]{1}", line) is None:
                print('Caught problem')
                continue
            mnem_end = re.search("[.]{1}", line).end()
            unit_end = mnem_end + re.search("[ ]{1}", line[mnem_end:]).end()
            colon_end = unit_end + re.search("[:]{1}", line[unit_end:]).start()

            # divide line
            mnem = line[:mnem_end - 1].strip()
            mnem = rename_depth(mnem)
            unit = line[mnem_end:unit_end].strip()
            data = line[unit_end:colon_end].strip()
            desc = line[colon_end + 1:].strip()
            # in some las file, the description contains the column number at the start
            # use a regex to find an initial number, and remove it
            test = re.findall(r"^\d+", desc)
            if len(test) > 0:
                desc = desc.replace(test[0], '')
                desc = desc.strip()

            # convert empty string ("") to None
            if len(data) == 0: data = None
            if section == "well_info" and mnem == "NULL":
                # save standard LAS NULL value
                null_val = data.rstrip('0')
                #data = None # this line seems strange and uncessary

            # parse data to type bool or number
            # BUT it also parsed well names as floats, which we should avoid
            if data is not None:
                if desc == 'WELL':
                    # avoid the __parse() function
                    pass
                elif data == "NO":
                    data = False
                elif data == "YES":
                    data = True
                else:
                    data = parse(data)

            # dynamic key
            key = "api_code" if section == "curve" else "value"
            content = {
                key: data,
                "unit": unit,
                "desc": desc
            }

            well_dict = add_section(well_dict, section, mnem.lower(), content)

        elif section == "data":
            content = line.split()
            for k, v in zip(generated_keys, content):
                #v = float(v) if v != null_val else None
                # replace all null values with np.nan, then we have a unified NaN in all well objects
                v = float(v) if v.rstrip('0') != null_val else np.nan
                well_dict = add_section(well_dict, section, k.lower(), v)

        elif section == "other":
            well_dict = add_section(well_dict, section, None, line)

    return null_val, generated_keys, well_dict
