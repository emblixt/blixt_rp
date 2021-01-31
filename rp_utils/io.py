import pandas as pd
import os
import re
import numpy as np
from datetime import datetime
from openpyxl import load_workbook, Workbook
from copy import deepcopy
import logging

from rp_utils.utils import isnan, info

logger = logging.getLogger(__name__)


def project_wells(filename, working_dir, all=False):
    """
    Returns a table containing the requested wells

    :param filename:
    :param working_dir:
    :param all:
        bool
        if True, all wells are loaded in the table, and not only the ones for which Use = Yes
    :return:
    """
    table = pd.read_excel(filename, header=1, sheet_name='Wells table')
    result = {}
    for i, ans in enumerate(table['Use']):
        # skip empty rows
        if not isinstance(ans, str):
            continue
        if all:
            ans = 'Yes'
        if ans.lower() == 'yes':
            temp_dict = {}
            log_dict = {}
            for key in list(table.keys()):
                if (key.lower() == 'las file') or (key.lower() == 'use'):
                    continue
                elif (key.lower() == 'given well name') or (key.lower() == 'note') or (key.lower() == 'translate log names'):
                    if isinstance(table[key][i], str):
                        if key.lower() == 'given well name':
                            value = fix_well_name(table[key][i])
                        else:
                            value = table[key][i]
                        temp_dict[key] = value
                    else:
                        temp_dict[key] = None  # avoid NaN
                else:
                    val = table[key][i]
                    if isnan(val):
                        continue
                    else:
                        this_list = make_clean_list(table[key][i], small_cap=True)
                        for log_name in this_list:
                            log_dict[log_name] = key
                temp_dict['logs'] = log_dict
            temp_file = test_file_path(table['las file'][i], working_dir)
            if temp_file is False:
                warn_txt = 'Warning, las file {} does not exist'.format(table['las file'][i])
                logger.warning(warn_txt)
                raise Warning(warn_txt)
            result[test_file_path(table['las file'][i], working_dir)] = temp_dict
    return result


def invert_well_table(well_table, well_name, rename=True):
    """
    Typically, the "log_table"
    :param well_table:
        dict
        As output from project_wells() above
    :param well_name:
        str
        name of the well we want to extract the "inverted well table" from well_table
    :param rename:
        bool
        if True it uses the "Translate log name" information to rename log names
    :return:
        dict
        As opposed to the commonly used "log_table", which relates a log type with one specific log, this dictionary
        relates a log type with multiple log names
        E.G. {'Resisitivity': ['rdep', 'rmed', 'rsha'], ...}
    """
    out = {}
    rdt = None
    if rename:
        rdt = get_rename_logs_dict(well_table)
    for key in list(well_table.keys()):
        if well_table[key]['Given well name'] == well_name:
            for lname, logtype in well_table[key]['logs'].items():
                _renamed = False
                if logtype not in list(out.keys()):
                    out[logtype] = []
                if rename and (rdt is not None):
                    for to_name, from_names in rdt.items():
                        if lname.lower() in from_names:
                            _renamed = True
                            out[logtype].append(to_name.lower())
                if not _renamed:
                    out[logtype].append(lname.lower())
    return out


def get_rename_logs_dict(well_table):
    """
    Interprets the "Translate log names"  keys of the well_table and returns a rename_logs dict.

    :param well_table:
        dict
        as returned from project_wells()
    :return:
        dict or None
    """
    rename_logs = {}
    for las_file, val in well_table.items():
        if val['Translate log names'] is None:
            continue
        _dict = interpret_rename_string(val['Translate log names'])
        for key in list(_dict.keys()):
            if key in list(rename_logs.keys()):
                if not _dict[key] in rename_logs[key]:  # only insert same rename pair once
                    rename_logs[key].append(_dict[key].lower())
            else:
                rename_logs[key] = [_dict[key].lower()]
    if len(rename_logs) < 1:
        return None
    else:
        return rename_logs


def project_templates(filename):
    table = pd.read_excel(filename, header=1, sheet_name='Templates')
    result = {}
    for i, ans in enumerate(table['Log type']):
        if not isinstance(ans, str):
            continue
        result[ans] = {}
        for key in ['bounds', 'center', 'colormap', 'description', 'max', 'min',
                    'scale', 'type', 'unit', 'line color', 'line style', 'line width']:
            result[ans][key] = None if isnan(table[key][i]) else table[key][i]
        result[ans]['full_name'] = ans

    # Also add the well settings
    table = pd.read_excel(filename, header=1, sheet_name='Well settings')
    for i, ans in enumerate(table['Given well name']):
        if not isinstance(ans, str):
            continue
        result[ans.upper()] = {}
        for key in ['Color', 'Symbol', 'Content', 'KB', 'UWI', 'UTM', 'X', 'Y', 'Water depth', 'Note']:
            result[ans.upper()][key.lower()] = None if isnan(table[key][i]) else table[key][i]

    return result


#def project_well_settings(filename):
#    table = pd.read_excel(filename, header=1, sheet_name='Well settings')
#    result = {}
#    for i, ans in enumerate(table['Given well name']):
#        result[ans] = {}
#        for key in ['Color', 'Symbol', 'Content', 'KB', 'UWI', 'UTM', 'X', 'Y', 'Water depth', 'Note']:
#            result[ans][key.lower()] = None if isnan(table[key][i]) else table[key][i]
#    return result


def project_working_intervals(filename):
    table = pd.read_excel(filename, header=4, sheet_name='Working intervals')
    result = {}
    return return_dict_from_tops(table, 'Given well name', 'Interval name', 'Top depth', include_base='Base depth')


def collect_project_wells(well_table, target_dir):
    """
    Copies all las files in the well table (the output from project_wells()) to the folder target_dir.

    :param well_table:
    :return:
    """
    from shutil import copyfile, SameFileError
    for las_file in list(well_table.keys()):
        if os.path.isfile(las_file):
            short_name = os.path.split(las_file)[-1]
            print('Copying file {} to {}'.format(
                short_name, target_dir
            ))
            try:
                copyfile(las_file, os.path.join(target_dir, short_name))
            except SameFileError:
                print('  File {} exists in target directory. Skipping to next'.format(short_name))
                continue


def read_sums_and_averages(filename, header=20):
    table = pd.read_excel(filename, header=header)
    unique_layers = unique_names(table, 'Name', well_names=False)
    answer = {}
    for layer in unique_layers:
        answer[layer] = {}

    for key in list(table.keys()):
        if key == 'Name':
            continue
        for i, value in enumerate(table[key]):
            answer[table['Name'][i]][key] = value

    return answer


def write_sums_and_averages(filename, line_of_data):
    # This function creates xlsx files. so please use Excel to save them as
    # xls files before attempting to load them into RokDoc
    # Additional columns are added after the  'ShaleVolumeAspectRatio' column, which are not
    # read by RokDoc
    if filename.split('.')[-1] == 'xls':
        filename += 'x'

    from openpyxl import load_workbook, Workbook
    if not os.path.isfile(filename):
        print('Creating new RokDoc Sums and Averages file')
        newfile = True
        wb = Workbook()
    else:
        print('Appending to existing RokDoc Sums and averages file')
        newfile = False
        wb = load_workbook(filename)

    ws = wb.active
    if newfile:
        ws.append(['Averages Set output from simple python script well_tops.py on {}'.format(
            datetime.now().isoformat())])
        ws.append(['Template Version: 1'])
        ws.append(['Depth units:             m'])
        ws.append(['Time units:              ms'])
        ws.append(['Velocity units:          m/s'])
        ws.append(['Density units:           g/cm3'])
        ws.append(['Porosity units:          fract'])
        ws.append(['AI units:                g/cm3_m/s'])
        ws.append(['SI units:                g/cm3_m/s'])
        ws.append(['M units:                 GPa'])
        ws.append(['MU units:                GPa'])
        ws.append(['K (Bulk Modulus) units:  GPa'])
        ws.append(['Lambda units:            GPa'])
        ws.append(['E units:                 GPa'])
        ws.append(['Lambda Mu units:         fract'])
        ws.append(['Mu Rho units:            GPa_g/cm3'])
        ws.append(['Lambda Rho units:        GPa_g/cm3'])
        ws.append(['Saturation units:        fract'])
        ws.append(['Volume units:            fract'])
        ws.append(['TableStart:'])
        ws.append(
            [
                'Name', 'Well', 'ZType', 'TopDepth', 'BaseDepth', 'MidPointDepth',
                'VpMean', 'VsMean', 'RhoMean', 'VpMedian', 'VsMedian', 'RhoMedian',
                'VpMode', 'VsMode', 'RhoMode', 'PorosityType', 'PorosityMean',
                'PorosityStdDev', 'Net', 'NetToGross', 'EpsilonMean', 'DeltaMean',
                'GammaMean', 'EpsilonMedian', 'DeltaMedian', 'GammaMedian',
                'EpsilonMode', 'DeltaMode', 'GammaMode', 'VpStdDev', 'VsStdDev',
                'RhoStdDev', 'EpsilonStdDev', 'DeltaStdDev', 'GammaStdDev',
                'VpVsCorrCoef', 'VpRhoCorrCoef', 'VsRhoCorrCoef', 'AI', 'SI',
                'M', 'MU', 'KBulkModulus', 'PR', 'Lambda', 'E', 'LambdaMu',
                'MuRho', 'LambdaRho', 'ShaleVolumeMean', 'ShaleVolumeStdDev',
                'ShaleVolumeInclusionShape', 'ShaleVolumeAspectRatio', 'Classification',
                'DateAdded'
            ]
        )
    ws.append(line_of_data)
    wb.save(filename)
    wb.close()


def read_tops(filename, top=True, zstick='md', frmt=None, only_these_wells=None):
    """

    :param filename:
    :param top:
    :param zstick:
    :param frmt:
    :param only_these_wells:
        list
        list of well names to look for, so that the reading in can be speeded up
        Populate this list by extracting a well_table, and use the well listed there
        > well_table = rp_utils.io.project_wells(project_table_file)
        > only_these_wells = list(set([x['Given well name'] for x in well_table.values()]))
        NOTE! The naming convention of the wells in the project table file must be the same as the one
        used in tops
    :return:
    """
    if frmt == 'petrel':
        return read_petrel_tops(filename, top=top, zstick=zstick, only_these_wells=only_these_wells)
    elif frmt == 'npd':
        return read_npd_tops(filename, top=top, zstick=zstick, only_these_wells=only_these_wells)
    elif frmt == 'rokdoc':
        return read_rokdoc_tops(filename, top=top, zstick=zstick, only_these_wells=only_these_wells)
    else:
        raise IOError('No tops for format {}'.format(frmt))


def read_rokdoc_tops(filename, header=4, top=True, zstick='md', only_these_wells=None):
    """
    :param top:
        bool
        if True, the top of each marker/top is returned
        if False, not implemented
    :param zstick:
        str
        adapted after RokDoc.
        Can be:
            'md', 'tvdkb','twt', 'tvdss',
    :param only_these_wells:
        list
        list of well names to look for, so that the reading in can be speeded up
    """
    if not top:
        raise NotImplementedError('Only top of markers / tops are available')

    if zstick == 'md':
        key_name = 'MD'
    elif zstick == 'tvdkb':
        key_name = 'TVDkb'
    elif zstick == 'twt':
        key_name = 'TWT'
    elif zstick == 'tvdss':
        key_name = 'TVDss'
    else:
        key_name = None
        NotImplementedError('ZSTICK = {} is not implemented'.format(zstick))

    tops = pd.read_excel(filename, header=header)
    return return_dict_from_tops(tops, 'Well Name', 'Horizon', key_name, only_these_wells=only_these_wells)


def read_npd_tops(filename, header=None, top=True, zstick='md', only_these_wells=None):
    """

    :param filename:
    :param header:
    :param top:
    :param zstick:
    :param only_these_wells:
        list
        list of well names to look for, so that the reading in can be speeded up
    :return:
    """
    if zstick != 'md':
        NotImplementedError('ZSTICK = {} is not implemented'.format(zstick))
    if top:
        key_name = 'Top depth [m]'
    else:
        key_name = 'Bottom depth [m]'

    tops = pd.read_excel(filename)
    return return_dict_from_tops(tops, 'Wellbore name', 'Lithostrat. unit', key_name, only_these_wells=only_these_wells)


def read_petrel_tops(filename, header=None, top=True, zstick='md', only_these_wells=None):
    if zstick != 'md':
        NotImplementedError('ZSTICK = {} is not implemented'.format(zstick))
    if not top:
        NotImplementedError('Only MD is implemented for Petrel top files')
        key_name = None
    else:
        key_name = 'MD'

    tops = pd.read_excel(filename)
    return return_dict_from_tops(tops, 'Well identifier', 'Surface', key_name, only_these_wells=only_these_wells)


def write_tops(filename, tops, well_names=None, interval_names=None, sheet_name=None):
    """
    Writes the tops to the excel file "filename", in the sheet name 'Working intervals'
    If "filename" exists, and is open, it raises a warning

    :param filename:
        str
        full pathname of excel file to write to.
        Assumes we're trying to write to the default project_table.xlsx, in the 'Working intervals' sheet.

    :param tops:
        dict
        As output from rp_utils.io.read_tops()
        {'well_A name': {'top1 name': top1_depth, 'top2 name': top2_depth, ...},  'well_B name': {...} }

    :param well_names:
        list
        list of str's
        list of names of the wells we would like to save to to file
        If None, all wells are saved

    :param interval_names:
        list
        list of str's
        list of names of the intervals (working intervals) we would like to save to to file,
        if None, all intervals are saved

    :return:
    """
    if sheet_name is None:
        sheet_name = 'Working intervals'

    # test write access
    taccs = check_if_excelfile_writable(filename)
    if not taccs:
        warn_txt = 'Not possible to write to {}'.format(filename)
        return

    if not os.path.isfile(filename):
        wb = Workbook()
    else:
        wb = load_workbook(filename)

    if sheet_name not in wb.sheetnames:
        print('Creating new sheet')
        ws = wb.create_sheet(sheet_name, -1)
    else:
        print('Opening existing sheet')
        ws = wb[sheet_name]

    # modify first line
    ws['A1'] = info()

    # test if fifth row exists
    if ws[5][0].value is None:
        ws['A2'] = 'Depth are in meters MD'
        for j, val in enumerate(['Use', 'Given well name', 'Interval name', 'Top depth', 'Base depth']):
            ws.cell(5, j+1).value = val

    # start appending data
    if well_names is None:
        well_names = list(tops.keys())

    for wname in well_names:
        these_tops = list(tops[wname].keys())
        if len(these_tops) == 0:
            continue  # skip wells without tops
        if interval_names is None:
            int_names = these_tops
            # Add a duplicate of the last interval to avoid running out-of-index
            int_names.append(int_names[-1])
        else:
            int_names = deepcopy(interval_names)

        # Add the 'TD' top name to catch it if it exists
        int_names.append('TD')

        # Find list of common top names
        ct = [tn for tn in int_names if tn in these_tops]
        # Find the index in these_tops to the last common top
        if len(ct) > 0:
            ind = these_tops.index(ct[-1])
            if ind + 1 <= len(these_tops):
                # if this is the last index of these_tops
                ct.append(ct[-1])
            else:
                # Add the next top in these_tops
                ct.append(these_tops[ind+1])
        else:
            ct.append(ct[-1])

        try:
            while 'TD' in ct:
                ct.remove('TD')
        except ValueError as ve:
            print(ve)


        for i in range(len(ct)-1):
            ws.append(['', wname, ct[i], tops[wname][ct[i]], tops[wname][ct[i+1]]])


    wb.save(filename)


def read_petrel_checkshots(filename, only_these_wells=None):
    checkshots = {}
    keys = []
    this_well_name = ''
    data_section = False
    header_section = False
    i = 0
    well_i = None

    with open(filename, 'r') as f:
        for line in f.readlines():
            if line[:7] == 'BEGIN H':
                header_section = True
                continue
            elif line[:5] == 'END H':
                header_section = False
                data_section = True
                continue
            if header_section:
                if line[:4].lower() == 'well':
                    well_i = i
                i += 1
                keys.append(line.strip())
            elif data_section:
                data = line.split()
                this_well_name = fix_well_name(data[well_i].replace('"', ''))
                if (only_these_wells is not None) and this_well_name not in only_these_wells:
                    continue
                if this_well_name not in list(checkshots.keys()):
                    checkshots[this_well_name] = {xx: [] for xx in keys}
                else:
                    for j, key in enumerate(keys):
                        checkshots[this_well_name][key].append(my_float(data[j]))

    return checkshots


def test_file_path(file_path, working_dir):
    # Convert backward slashes to forward slashes
    file_path = file_path.replace("\\", "/")
    if os.path.isfile(file_path):
        return file_path
    elif os.path.isfile(os.path.join(working_dir, file_path)):
        return os.path.join(working_dir, file_path)
    else:
        return False


def check_if_excelfile_writable(fnm):
    from openpyxl.utils.exceptions import InvalidFileException
    if os.path.exists(fnm):
        # path exists
        if os.path.isfile(fnm): # is it a file or a dir?
            # also works when file is a link and the target is writable
            # try to open and save it
            try:
                wb = load_workbook(fnm)
            except InvalidFileException as ie:
                print(ie)
                return False
            try:
                wb.save(fnm)
            except PermissionError as pe:
                print(pe)
                print('File is open. Please close it and try again')
                return False
            return True
        else:
            return False # path is a dir, so cannot write as a file

    # target does not exist, check perms on parent dir
    pdir = os.path.dirname(fnm)
    if not pdir:
        pdir = '.'
    # target is creatable if parent dir is writable
    return os.access(pdir, os.W_OK)


def fix_well_name(well_name):
    if isinstance(well_name, str):
        return well_name.replace('/', '_').replace('-', '_').replace(' ', '').upper()
    else:
        return


def unique_names(table, column_name, well_names=True):
    """
    :param table:
        panda object
        as returned from pandas.read_excel()
    returns the list of unique values in a column named 'column_name'
    """
    if well_names:
        return [fix_well_name(x) for x in list(set(table[column_name])) if isinstance(x, str)]
    else:
        return [x for x in list(set(table[column_name])) if isinstance(x, str)]


def return_dict_from_tops(tops, well_key, top_key, key_name, only_these_wells=None, include_base=None):
    """

    :param tops:
    :param well_key:
        str
        Name of the column which contain the well names
    :param top_key:
        str
        Name of the column which contain the tops / interval names
    :param key_name:
        str
        Name of the column which contain the top depth
    :param only_these_wells:
    :param include_base:
        str
        Name of the column which contain the base depth
        if this is used each top / interval will contain a list of of top and base depth
        else it is just the top depth
    :return:
    """
    if only_these_wells:
        unique_wells = only_these_wells
    else:
        unique_wells = unique_names(tops, well_key)
    answer = {}

    for well_name in unique_wells:
        answer[well_name] = {}

    for well_name in unique_wells:
        for i, marker_name in enumerate(list(tops[top_key])):
            if fix_well_name(tops[well_key][i]) != well_name:
                continue  # not on the right well
            if include_base is not None:
                answer[well_name][marker_name.upper()] = [tops[key_name][i], tops[include_base][i]]
            else:
                answer[well_name][marker_name.upper()] = tops[key_name][i]

    return answer


def make_clean_list(input_str, small_cap=False):
    if not isinstance(input_str, str):
        raise IOError('Only accept strings')
    if small_cap:
        return [x.strip().lower() for x in input_str.split(',') if x.strip() != '']
    else:
        return [x.strip() for x in input_str.split(',') if x.strip() != '']


def write_las(filename, wh, lh, data, overwrite=False):
    """
    Write the well data to a las format file.

    :param filename:
        str
        full path name to las file
    :param wh:
        core.well.Header
        well header =  w.header, where w is a core.well.Well object
    :param lh:
        core.well.Header
        log header = w.block['Block name'].header, where w is a core.well.Well object
    :param data:
        dict
        data = w.block['Block name'].logs, where w is a core.well.Well object
    :param overwrite:
        bool
        Set to True to allow overwriting an existing las file
    :return:
    """
    if os.path.isfile(filename) and (not overwrite):
        warn_txt = 'File {} already exist. Write cancelled'.format(filename)
        print('WARNING: {}'.format(warn_txt))
        logger.warning(warn_txt)
        return

    out = (
        '#----------------------------------------------------------------------------\n'
        '~VERSION INFORMATION\n'
        'VERS.            2.0                  :CWLS LOG ASCII STANDARD -VERSION 2.0\n'
        'WRAP.            NO                   :ONE LINE PER DEPTH STEP\n'
        '#\n'
    )

    out += '# {}\n'.format(wh['creation_info'].value)
    if 'note' in list(wh.keys()):
        out += '# NOTE: {}\n'.format(wh['note'].value)
    out += '# Written to las on: {}\n'.format(datetime.now().isoformat())
    out += '# Modified on: {}\n'.format(wh['modification_date'].value)
    for key, value in data.items():
        if value.header['modification_history'] is not None:
            out += '#  Modification: {}: {}\n'.format(key, value.header['modification_history'].replace('\n','\n#   '))

    out += (
        '#--------------------------------------------------------------------\n'
        '~WELL INFORMATION\n'
        '#MNEM .UNIT      DATA                 :DESCRIPTION OF MNEMONIC\n'
        '#----------      ------------         -------------------------------\n'
    )

    # add info about start stop etc. from Block header
    for key in list(lh.keys()):
        if key in ['name', 'creation_info', 'creation_date', 'modification_date', 'well']:
            continue
        out += '{0: <7}.{1: <9}{2: <21}:{3:}\n'.format(
            key.upper(),
            lh[key].unit,
            str(lh[key].value) if lh[key].value is not None else '',
            lh[key].desc.upper()
        )

    # add info well header
    for key in list(wh.keys()):
        if key in ['name', 'note', 'creation_info', 'creation_date', 'modification_date']:
            continue
        out += '{0: <7}.{1: <9}{2: <21}:{3:}\n'.format(
            key.upper(),
            wh[key].unit,
            str(wh[key].value) if wh[key].value is not None else '',
            wh[key].desc.upper()
        )

    out += (
        '#\n'
        '# ----------------------------------------------------------------------------\n'
        '~CURVE INFORMATION\n'
        '# MNEM.UNIT                                         : CURVE DESCRIPTION\n'
        '# ----------                                        -------------------------------\n'
    )
    i = 1
    out += '{0: <20}.{1: <33}: {2: <9}{3:}\n'.format(
        'DEPTH',
        data['depth'].header['unit'],
        i,
        data['depth'].header['desc']
    )
    for key, value in data.items():
        if key == 'depth':
            continue
        i += 1
        out += '{0: <20}.{1: <33}: {2: <9}{3:}\n'.format(
            key.upper(),
            value.header['unit'],
            i,
            value.header['desc']
        )
    out += (
        '#\n'
        '# ----------------------------------------------------------------------------\n'
        '~A                  '
    )

    # write data column headers
    for key in list(data.keys()):
        if key == 'depth':
            continue
        out += '{0: <20}'.format(key.upper())
    out += '\n'

    # start writing data
    for i, md in enumerate(data['depth'].data):
        out += '{0: <20}'.format(md)
        for key in list(data.keys()):
            if key == 'depth':
                continue
            out += '{0: <20}'.format(
                wh['null'].value if np.isnan(data[key].data[i]) else data[key].data[i]
            )
        out += '\n'


    with open(filename, 'w+') as f:
        f.write(out)


def get_las_header(filename):
    """
    Iterates over the las files header lines.

    :param filename:
    :return:
    """
    for row in open(filename, 'r'):
        if '~W' in row:
            break
        else:
            yield row


def get_las_well_info(filename):
    """
    Iterates over the las file well info lines.

    :param filename:
    :return:
    """
    well_info_section = False
    curve_info_section = False
    header = True
    for row in open(filename, 'r'):
        # test wich section you are in
        if '~W' in row:
            well_info_section = True
        if '~C' in row:
            curve_info_section = True

        if curve_info_section:
            break
        if well_info_section:
            header = False
            yield row
        if header:
            continue


def get_las_curve_info(filename):
    """
    Iterates over the las file curve info lines.

    :param filename:
    :return:
    """
    curve_info_section = False
    data_section = False
    header = True
    for row in open(filename, 'r'):
        # test wich section you are in
        if '~C' in row:
            curve_info_section = True
        if '~A' in row:
            data_section = True

        if data_section:
            break
        if curve_info_section:
            header = False
            yield row
        if header:
            continue


def convert(lines, file_format='las'):
    """
    class handling wells, with logs, and well related information
    The reading .las files is more or less copied from converter.py
        https://pypi.org/project/las-converter/

    """
    def parse(x):
        try:
            x = int(x)
        except ValueError:
            try:
                x = float(x)
            except ValueError:
                pass
        return x

    def get_current_section(line):
        if '~V' in line : return 'version'
        if '~W' in line: return 'well_info'
        if '~C' in line: return 'curve'
        if '~P' in line: return 'parameter'
        if '~O' in line: return 'other'
        if '~A' in line: return 'data'
        # ~ Unregistered section
        return None

    def add_section(well_dict, section, mnem, content):
        if section == "data":
            if isinstance(content, list):
                well_dict[section][mnem] = content
            else:
                well_dict[section][mnem].append(content)
        elif section == "other":
            well_dict[section] += "".join(str(content).strip())
        else:
            well_dict[section][mnem] = content
        return well_dict

    generated_keys = []
    null_val = None
    section = ""
    length_units = ['m', 'ft']
    rules = {"version", "well_info", "parameter", "curve"}
    descriptions = []
    curve_names = None
    well_dict = {"version": {}, "well_info": {}, "curve": {}, "parameter": {}, "data": {}, "other": ""}
    if file_format == 'RP well table':
        null_val = 'NaN'

    for line in lines:
        content = {}

        if isinstance(line, bytes):
            line = line.decode("utf-8").strip()

        # line just enter or "\n"
        if len(line) <= 1: continue
        # comment
        if "#" in line: continue

        # section
        if "~" in line:
            section = get_current_section(line)

            # get section version first
            if section == "version":
                continue

            # generate keys of log[data] based on log[curve]
            if section == "data":
                generated_keys = [e.lower() for e in well_dict["curve"].keys()]
                for key in generated_keys:
                    # XXX
                    #key = rename_log_name(key)
                    # inital all key to empty list
                    well_dict = add_section(well_dict, section, key, [])

            continue

        if file_format == 'RP well table':
            if line[:7] == 'Columns':
                section = 'RP header'
                continue  # jump into header

            if section == 'RP header':
                if line[2:5] == ' - ':
                    descriptions.append(line.split(' - ')[-1].strip())

            if line[:7] == 'Well ID':
                # parse curve names
                curve_names = [t.strip().lower() for t in line.split('\t')]
                section = 'dummy_value'

            if line[:4] == '  No':
                # parse line of units
                # unit_names = [t.strip() for t in line.split('\t')]
                unit_names = [t.strip() for t in line.split()]
                unit_names = [t.replace('[', '').replace(']', '') for t in unit_names]

                for this_curve_name, this_unit_name, this_description in zip(curve_names, unit_names, descriptions):
                    well_dict = add_section(well_dict, 'curve',
                                            this_curve_name,
                                            {'api_code': None, 'unit': this_unit_name.lower(), 'desc': this_description}
                                            )
                generated_keys = [key for key in curve_names]
                section = 'data'
                # initiate all key to empty list
                for key in generated_keys:
                    # TODO This will probably fail. Test and compare with 'original' converter
                    well_dict.add_section(well_dict, key, [])
                continue  # jump into data

        # unregistered section
        if section is None: continue

        if section in rules:
            # index of seperator
            if re.search("[.]{1}", line) is None:
                print('Caught problem')
                continue
            mnem_end = re.search("[.]{1}", line).end()
            unit_end = mnem_end + re.search("[ ]{1}", line[mnem_end:]).end()
            colon_end = unit_end + re.search("[:]{1}", line[unit_end:]).start()

            # divide line
            mnem = line[:mnem_end - 1].strip()
            # XXX
            # mnem = rename_log_name(mnem)
            unit = line[mnem_end:unit_end].strip().lower()
            data = line[unit_end:colon_end].strip()
            # in some las files, the unit is given directly behind the data, e.g "30.0 M"
            # When this is the case, clean the data and try add the unit to the unit
            if (data[-2:].lower().strip() in length_units) or (data[-3:].lower().strip() in length_units):
                _tmp = data.split()
                data = _tmp[0].strip()
                _unit = _tmp[-1].strip().lower()
                if len(unit) == 0:
                    unit = _unit
            desc = line[colon_end + 1:].strip()

            # in some las file, the description contains the column number at the start
            # use a regex to find an initial number, and remove it
            test = re.findall(r"^\d+", desc)
            if len(test) > 0:
                desc = desc.replace(test[0], '')
                desc = desc.strip()

            # convert empty string ("") to None
            if len(data) == 0: data = None
            if section == "well_info" and mnem == "NULL":
                # save standard LAS NULL value
                null_val = data.rstrip('0')
                #data = None # this line seems strange and uncessary

            # parse data to type bool or number
            # BUT it also parsed well names as floats, which we should avoid
            if data is not None:
                if desc == 'WELL' or mnem == 'WELL' or mnem == 'UWI':
                    #if section == "well_info" and (mnem == "WELL" or mnem == 'UWI'):
                    # catch well name, harmonize it, and avoid the parse() function
                    data = fix_well_name(data)
                elif data == "NO":
                    data = False
                elif data == "YES":
                    data = True
                else:
                    data = parse(data)

            # dynamic key
            key = "api_code" if section == "curve" else "value"
            content = {
                key: data,
                "unit": unit,
                "desc": desc
            }

            well_dict = add_section(well_dict, section, mnem.lower(), content)

        elif section == "data":
            content = line.split()
            for k, v in zip(generated_keys, content):
                #v = float(v) if v != null_val else None
                # replace all null values with np.nan, then we have a unified NaN in all well objects
                v = float(v) if v.rstrip('0') != null_val else np.nan
                well_dict = add_section(well_dict, section, k.lower(), v)

        elif section == "other":
            well_dict = add_section(well_dict, section, None, line)

    return null_val, generated_keys, well_dict


def interpret_rename_string(rename_string):
    """
    creates a rename dictionary ({'VCL': 'VSH', 'Vp': 'Vp_dry'}) from input string
    :param rename_string:
        str
        renaming defined by "VSH->VCL, Vp_dry->Vp"
    :return:
        dict or None
    """
    if len(rename_string) < 3:
        return None

    return_dict = {}
    for pair in rename_string.split(','):
        if '->' not in pair:
            continue
        names = pair.split('->')
        if len(names) > 2:
            warn_txt = "Translation pairs should be separated by ',': ".format(pair)
            print('WARNING: {}'.format(warn_txt))
            logger.warning(warn_txt)
            continue
        return_dict[names[1].strip().lower()] = names[0].strip().lower()
    if len(return_dict) < 1:
        return None
    else:
        return return_dict

def interpret_cutoffs_string(cutoffs_string):
    """
    creates a cutoffs dictionary ({'VCL': ['<', 0.5], 'PHIE': ['>', 0.1]}) from input string
    :param cutoffs_string:
        str
        renaming defined by "VCL>0.8, PHIE>0.1"
    :return:
        dict or None
    """
    if len(cutoffs_string) < 3:
        return None

    return_dict = {}
    for pair in cutoffs_string.split(','):
        m_symb = None
        # search for masking symbol
        if '==' in pair:
            m_symb = '=='  # equal
        elif '<=' in pair:
            m_symb = '<='  # less or equal
        elif '>=' in pair:
            m_symb = '>='  # greater or equal
        elif '!=' in pair:
            m_symb = '!='  # not equal
        elif '>' in pair:
            m_symb = '>'  # greater
        elif '<' in pair:
            m_symb = '<'  # less
        else:
            warn_txt = 'No valid masking symbol given in {}'.format(pair)
            print('WARNING: {}'.format(warn_txt))
            logger.warning(warn_txt)
            continue

        tmp = pair.split(m_symb)
        if len(tmp) > 2:
            warn_txt = "Something fishy in cutoffs string: ".format(cutoffs_string)
            print('WARNING: {}'.format(warn_txt))
            logger.warning(warn_txt)
            continue
        return_dict[tmp[0].strip().lower()] = [m_symb, float(tmp[1].strip())]
    if len(return_dict) < 1:
        return None
    else:
        return return_dict

def my_float(string):
    try:
        return float(string)
    except ValueError:
        return string

def rename_log_name(_key):
    """
    Helper function that translates different "depth" names to a common "depth" name.
    :param _key:
        str
    :return:
        str
    """
    # This function is stalled, and will not work right now\
    rename_well_logs = {}
    for rname, value in rename_well_logs.items():
        if _key.lower() in [x.lower() for x in value]:
            info_txt = 'Renaming log from {} to {}'.format(_key, rname)
            print('INFO: {}'.format(info_txt))
            logger.info(info_txt)
            return rname.lower()
    else:
        return _key


