# -*- coding: utf-8 -*-
"""
Created on Fri Nov 22 13:30:57 2019
@author: mblixt
"""
import pandas as pd
import sys
import os
import numpy as np
from datetime import datetime
import matplotlib.pyplot as plt
import matplotlib.colors as colors


import masks as msks


plt.ioff()

msymbols = np.array(['o','s','v','^','<','>','p','*','h','H','+','x','D','d','|','_','.','1','2','3','4','8'])
cnames = list(np.roll([str(u) for u in colors.cnames.keys()], 10))


def fix_well_name(well_name):
    if isinstance(well_name, str):
        return well_name.replace('/', '_').replace('-', '_').replace(' ', '').upper()
    else:
        return


def unique_names(table, column_name, well_names=True):
    """
    :param table:
        panda object
        as returned from pandas.read_excel()

    returns the list of unique values in a column named 'column_name'
    """
    if well_names:
        return [fix_well_name(x) for x in list(set(table[column_name])) if isinstance(x, str)]
    else:
        return [x for x in list(set(table[column_name])) if isinstance(x, str)]


def return_dict_from_tops(tops, well_key, top_key, key_name):
    unique_wells = unique_names(tops, well_key)
    answer = {}
    
    for well_name in unique_wells:
        answer[well_name] = {}
    
    for well_name in unique_wells:
        for i, marker_name in enumerate(list(tops[top_key])):
            if fix_well_name(tops[well_key][i]) != well_name:
                continue  # not on the right well
            answer[well_name][marker_name.upper()] = tops[key_name][i]
    
    return answer


def get_translated_name(log_curve_names, this_type='depth'):
    # Finds the curve name among the curve names that most likely correspond to
    # desired curve
    translated_name = None
    tentative_names = None

    if this_type == 'depth':
        tentative_names = ['dept', 'depth']
    else:
        return this_type

    for tentative_name in tentative_names:
        if tentative_name in log_curve_names:
            translated_name = tentative_name

    return translated_name


def read_sums_and_averages(filename, header=20):
    table = pd.read_excel(filename, header=header)
    unique_layers = unique_names(table, 'Name', well_names=False)
    answer = {}
    for layer in unique_layers:
        answer[layer] = {}

    for key in list(table.keys()):
        if key == 'Name':
            continue
        for i, value in enumerate(table[key]):
            answer[table['Name'][i]][key] = value

    return answer



def write_sums_and_averages(filename, line_of_data):
    # This function creates xlsx files. so please use Excel to save them as 
    # xls files before attempting to load them into RokDoc
    if filename.split('.')[-1] == 'xls':
        filename += 'x'
    
    from openpyxl import load_workbook, Workbook
    if not os.path.isfile(filename):
        print('Creating new RokDoc Sums and Averages file')
        newfile = True
        wb = Workbook()
    else:
        print('Appending to existing RokDoc Sums and averages file')
        newfile = False
        wb = load_workbook(filename)

    ws = wb.active
    if newfile:
        ws.append(['Averages Set output from simple python script well_tops.py on {}'.format(
                datetime.now().isoformat())])
        ws.append(['Template Version: 1'])
        ws.append(['Depth units:             m'])
        ws.append(['Time units:              ms'])
        ws.append(['Velocity units:          m/s'])
        ws.append(['Density units:           g/cm3'])
        ws.append(['Porosity units:          fract'])
        ws.append(['AI units:                g/cm3_m/s'])
        ws.append(['SI units:                g/cm3_m/s'])
        ws.append(['M units:                 GPa'])
        ws.append(['MU units:                GPa'])
        ws.append(['K (Bulk Modulus) units:  GPa'])
        ws.append(['Lambda units:            GPa'])
        ws.append(['E units:                 GPa'])
        ws.append(['Lambda Mu units:         fract'])
        ws.append(['Mu Rho units:            GPa_g/cm3'])
        ws.append(['Lambda Rho units:        GPa_g/cm3'])
        ws.append(['Saturation units:        fract'])
        ws.append(['Volume units:            fract'])
        ws.append(['TableStart:'])
        ws.append(
                [
                'Name', 'Well', 'ZType', 'TopDepth', 'BaseDepth', 'MidPointDepth',	
                'VpMean', 'VsMean', 'RhoMean', 'VpMedian', 'VsMedian', 'RhoMedian', 
                'VpMode', 'VsMode', 'RhoMode', 'PorosityType', 'PorosityMean', 
                'PorosityStdDev', 'Net', 'NetToGross', 'EpsilonMean', 'DeltaMean', 
                'GammaMean', 'EpsilonMedian', 'DeltaMedian', 'GammaMedian', 
                'EpsilonMode', 'DeltaMode', 'GammaMode', 'VpStdDev', 'VsStdDev', 
                'RhoStdDev', 'EpsilonStdDev', 'DeltaStdDev', 'GammaStdDev', 
                'VpVsCorrCoef', 'VpRhoCorrCoef', 'VsRhoCorrCoef', 'AI', 'SI', 
                'M', 'MU', 'KBulkModulus', 'PR', 'Lambda', 'E', 'LambdaMu', 
                'MuRho', 'LambdaRho', 'ShaleVolumeMean', 'ShaleVolumeStdDev', 
                'ShaleVolumeInclusionShape', 'ShaleVolumeAspectRatio'
                ]
        )
    ws.append(line_of_data)
    wb.save(filename)
    wb.close()


def read_rokdoc_tops(filename, header=4, top=True, zstick='md'):
    """
    :param top:
        bool
        if True, the top of each marker/top is returned
        if False, not implemented
    
    :param zstick:
        str
        adapted after RokDoc. 
        Can be:
            'md', 'tvdkb','twt', 'tvdss', 
    
    """
    if not top:
        raise NotImplementedError('Only top of markers / tops are available')
    
    if zstick == 'md': key_name = 'MD'
    elif zstick == 'tvdkb': key_name = 'TVDkb'
    elif zstick == 'twt': key_name = 'TWT'
    elif zstick == 'tvdss': key_name = 'TVDss'  
    else:
        key_name = None
        NotImplementedError('ZSTICK = {} is not implemented'.format(zstick))

    tops = pd.read_excel(filename, header=header)
    return return_dict_from_tops(tops, 'Well Name', 'Horizon', key_name)


def read_npd_tops(filename, header=None, top=True, zstick='md'):
    if zstick != 'md':
        NotImplementedError('ZSTICK = {} is not implemented'.format(zstick))
    if top:
        key_name = 'Top depth [m]'
    else:
        key_name = 'Bottom depth [m]'

    tops = pd.read_excel(filename)
    return return_dict_from_tops(tops, 'Wellbore name', 'Lithostrat. unit', key_name)


def read_petrel_tops(filename, header=None, top=True, zstick='md'):
    if zstick != 'md':
        NotImplementedError('ZSTICK = {} is not implemented'.format(zstick))
    if not top:
        NotImplementedError('Only MD is implemented for Petrel top files')
        key_name = None
    else:
        key_name = 'MD'
        
    tops = pd.read_excel(filename)
    return return_dict_from_tops(tops, 'Well identifier', 'Surface', key_name)


def nan_corrcoef(x,y):
    maskx = ~np.ma.masked_invalid(x).mask
    masky = ~np.ma.masked_invalid(y).mask
    mask = msks.combine_masks([maskx, masky])
    return np.corrcoef(x[mask], y[mask])


def calc_stats(
        tops,
        well_files,
        intervals,
        logs,
        cutoffs,
        save_to=None,
        suffix=None
):
    """
    Loop of over a set of wells, and a well tops dictionary and calculate the statistics over all wells within
    specified intervals.
    :param tops:
        dict
        E.G.
        tops_file = 'C:/Users/marten/Google Drive/Blixt Geo AS/Projects/Well log plotter/Tops/strat_litho_wellbore.xlsx'
        tops = read_npd_tops(tops_file)
    :param well_files:
        list
        list of filenames pointing to las files that we want to examinate
        E.G.
            well_files = [fbase + xx for xx in [
               '7321_7-1_CPI.las',
               '7321_8-1_CPI.las']
    :param intervals:
        list of dicts
        E.G.
            [
                {'name': 'Hekkingen sands',
                 'tops': ['HEKKINGEN FM', 'BASE HEKKINGEN']},
                {'name': 'Kolmule sands',
                 'tops': ['KOLMULE FM', 'BASE KOLMULE'
            ]
        The 'name' of the interval is used in the saved RokDoc compatible sums and averages file
        to name the averages
    :param logs:
        dict
        Dictionary of log names to create statistics on
        The Vp, Vs, Rho and Phi logs are necessary for output to RokDoc compatible Sums & Average excel file
        E.G.
            logs = {'vp': 'vp', 'vs': 'vs', 'rho': 'rhob', 'phi': 'phie', 'vcl': 'vcl'}
        Must have the same name across all las files. And for proper export to RokDoc Sums and Average files, they have
        have to have these names
    :param cutoffs:
        dict
        Dictionary with log names as keys, and list with mask definition as value
        E.G.
            {'vcl': ['<', 0.5], 'phie': ['>', 0.1]}
    :param save_to:
        str
        name of file of which should contain the averages (RokDoc format)
        and the base of the file name is where the plots are saved
    :param suffix:
        str
        Suffix added to output plots (png) to ease separating output from eachother
    :return:
    """
    import converter
    c = converter.Converter()
    
    if suffix is None:
        suffix = ''
    else:
        suffix = '_' + suffix
    
    cutoffs_str = ''
    for key in cutoffs:
        cutoffs_str += '{}{}{:.2f}, '.format(key, cutoffs[key][0], cutoffs[key][1])
    cutoffs_str = cutoffs_str.rstrip(', ')
        
    if save_to is not None:
        fbase = os.path.split(save_to)[0]
    else:
        fbase = None
    
    # open a figure for each log that is being analyzed
    figs_and_axes = [plt.subplots(figsize=(8,6)) for xx in range(len(logs))]
    interval_axis = []
    interval_ticks = ['', ]

    for j, interval in enumerate(intervals):
        interval_axis.append(j)
        interval_ticks.append(interval['name'])
        
        # create container for results
        results = {}
        results_per_well = {}
        for key in list(logs.values()):
            results[key.lower()] = np.empty(0)

        wells = [c.set_file(ff) for ff in well_files]
        depth_from_top = {}

        for well in wells:
            this_well_name = fix_well_name(well.well['well']['value'])
            results_per_well[this_well_name] = {}
            try:
                print('Well: {}'.format(this_well_name))
                print(' Top: {}: {:.2f} [m] MD'.format(interval['tops'][0], tops[this_well_name][interval['tops'][0].upper()]))
                print(' Base: {}: {:.2f} [m] MD'.format(interval['tops'][1], tops[this_well_name][interval['tops'][1].upper()]))
            except:
                print('Tops {} & {} not present in {}'.format(interval['tops'][0], interval['tops'][1], this_well_name))
                depth_from_top[this_well_name] = np.empty(0)
                for key in list(logs.values()):
                    results_per_well[this_well_name][key.lower()] = np.empty(0)
                continue
                

            # create masks based on cutoffs
            masks = []
            for key in list(cutoffs.keys()):
                this_data = np.array(well.data[key.lower()], dtype='float')
                masks.append(
                        msks.create_mask(this_data,
                                         cutoffs[key][0],
                                         cutoffs[key][1])
                )
            # create mask based on tops
            depth_key = get_translated_name(list(well.data.keys()), 'depth')
            this_depth = np.array(well.data[depth_key], dtype='float')
            masks.append(
                msks.create_mask(this_depth,
                     '><',
                     [
                             tops[this_well_name][interval['tops'][0].upper()],
                             tops[this_well_name][interval['tops'][1].upper()]
                     ]
                )
            )
            print('   Depth range: {:.1f} - {:.1f}'.format(
                                                            np.nanmin(this_depth[masks[-1]]),
                                                            np.nanmax(this_depth[masks[-1]])
                                                          )
            )
            # combine all masks
            this_mask = msks.combine_masks(masks)

            # calculate the depth from the top for each well
            depth_from_top[this_well_name] = this_depth[this_mask] - tops[this_well_name][interval['tops'][0].upper()]

            for key in list(logs.values()):
                key = key.lower()
                this_data = np.array(well.data[key], dtype='float')
                print('  {}:'.format(key))
                results[key] = np.append(results[key], this_data[this_mask])
                results_per_well[this_well_name][key] = this_data[this_mask]

        # create plot of logs vs depth and fill the interval plots
        ncols = len(logs)
        fig, axs = plt.subplots(
                nrows=1, 
                ncols=ncols, 
                sharey=True,
                figsize=(ncols*6, 8))
        for i, key in enumerate(logs.values()):
            key = key.lower()
            well_names = []
            for k, well in enumerate(wells):
                this_well_name = fix_well_name(well.well['well']['value'])
                well_names.append(this_well_name)
                axs[i].plot(
                        results_per_well[this_well_name][key],
                        depth_from_top[this_well_name],
                        c=cnames[k])
                
                mn =  np.nanmean(results_per_well[this_well_name][key]) 
                std = np.nanstd(results_per_well[this_well_name][key])
                figs_and_axes[i][1].errorbar(
                        mn, j, xerr=std, fmt='none', capsize=10,
                        capthick=1, elinewidth=1, ecolor=cnames[k])
                figs_and_axes[i][1].scatter(
                        mn, j, 
                        marker=msymbols[k], 
                        c=cnames[k],
                        s=90)
            
            axs[i].set_xlabel(key)
            axs[i].set_ylim(axs[i].get_ylim()[1], axs[i].get_ylim()[0])

            mn =  np.nanmean(results[key]) 
            std = np.nanstd(results[key])
            figs_and_axes[i][1].errorbar(
                    mn, j, xerr=std, fmt='none', capsize=20,
                    capthick=3, elinewidth=3, ecolor='pink')
            figs_and_axes[i][1].scatter(mn, j, marker='*', s=360, c='pink')            
            figs_and_axes[i][1].set_xlabel(key)
        
        axs[0].set_ylabel('Depth from {} [m]'.format(interval['tops'][0]))
        axs[-1].legend(well_names, prop={'size': 10})
        fig.suptitle('{}, {}'.format(interval['name'], cutoffs_str))
        fig.tight_layout()
        if fbase is not None:
            fig.savefig(os.path.join(fbase, '{}_logs_depth{}.png'.format(interval['name'], suffix)))
        
        # create histogram plot
        fig, axs = plt.subplots(nrows=ncols, ncols=1,figsize=(9, 8*ncols))
        for i, key in enumerate(logs.values()):
            key = key.lower()
            n, bins, patches = axs[i].hist(
                    [results_per_well[wid][key] for wid in well_names],
                    10, 
                    histtype='bar', 
                    stacked=True, 
                    label=well_names,
                    color=[cnames[k] for k in range(len(wells))]
            )
            axs[i].set_ylabel('N'); axs[i].set_xlabel(key); ylim = axs[i].get_ylim()
            mn =  np.nanmean(results[key]); std = np.nanstd(results[key])
            axs[i].plot([mn, mn], ylim, c='black', lw=2)
            axs[i].plot([mn+std, mn+std], ylim, 'b--')
            axs[i].plot([mn-std, mn-std], ylim, 'b--')
        axs[-1].legend(prop={'size': 10})
        axs[0].set_title('{}, {}'.format(interval['name'], cutoffs_str))
        fig.tight_layout()
        if fbase is not None:
            fig.savefig(os.path.join(fbase, '{}_logs_hist{}.png'.format(interval['name'], suffix)))
        
        # Write result to RokDoc compatible excel Sums And Average xls file:
        if save_to is not None:
            write_sums_and_averages(save_to,
                [
                     interval['name'],
                     'NONE',
                     'MD',
                     -999.25,
                     -999.25,
                     -999.25,
                     np.nanmean(results[logs['vp'].lower()]),
                     np.nanmean(results[logs['vs'].lower()]),
                     np.nanmean(results[logs['rho'].lower()]),
                     np.nanmedian(results[logs['vp'].lower()]),
                     np.nanmedian(results[logs['vs'].lower()]),
                     np.nanmedian(results[logs['rho'].lower()]),
                     np.nanmean(results[logs['vp'].lower()]),  # this should be a mode value, but uncertain what it means
                     np.nanmean(results[logs['vs'].lower()]),  # this should be a mode value, but uncertain what it means
                     np.nanmean(results[logs['rho'].lower()]),  # this should be a mode value, but uncertain what it means
                     'NONE',
                     np.nanmean(results[logs['phi'].lower()]),
                     np.nanstd(results[logs['phi'].lower()]),
                     -999.25,
                     -999.25,
                     -999.25,
                     -999.25,
                     -999.25,
                     -999.25,
                     -999.25,
                     -999.25,
                     -999.25,
                     -999.25,
                     -999.25,
                     np.nanstd(results[logs['vp'].lower()]),
                     np.nanstd(results[logs['vs'].lower()]),
                     np.nanstd(results[logs['rho'].lower()]),
                     -999.25,
                     -999.25,
                     -999.25,
                     nan_corrcoef(results[logs['vp'].lower()], results[logs['vs'].lower()])[0,1],
                     nan_corrcoef(results[logs['vp'].lower()], results[logs['rho'].lower()])[0,1],
                     nan_corrcoef(results[logs['vs'].lower()], results[logs['rho'].lower()])[0,1],
                     -999.25,
                     -999.25,
                     -999.25,
                     -999.25,
                     -999.25,
                     -999.25,
                     -999.25,
                     -999.25,
                     -999.25,
                     -999.25,
                     -999.25,
                     np.nanmean(results[logs['vcl'].lower()]),
                     np.nanstd(results[logs['vcl'].lower()]),
                     'None',
                     0.0
                        ]
            )
    # arrange the interval plots
    well_names.append('All')
    for i, fax in enumerate(figs_and_axes):
        fax[1].legend(well_names, prop={'size': 10})
        fax[1].grid(True)
        fax[1].set_title(cutoffs_str)
        fax[1].set_yticklabels(interval_ticks)
        fax[0].tight_layout()
        if fbase is not None:
            fax[0].savefig(os.path.join(
                    fbase,
                    '{}_{}_intervals{}.png'.format(list(logs.values())[i],
                            cutoffs_str.replace('>','_gt_').replace('<','_lt_').replace('=','_eq_'),
                            suffix)))
            plt.close('all')